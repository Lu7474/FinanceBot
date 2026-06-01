"""Excel export, backup, template generation and import parsing."""

from datetime import date as date_type
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import MAX_ACCOUNT_NAME_LENGTH, MAX_AMOUNT, MAX_CATEGORY_LENGTH
from core.utils import clean_text

# ── Style constants ───────────────────────────────────────────────────────────
_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_SECTION_FILL = PatternFill("solid", fgColor="CFE2F3")
_INCOME_FILL = PatternFill("solid", fgColor="D9EAD3")
_EXPENSE_FILL = PatternFill("solid", fgColor="FCE5CD")
_TOTAL_FILL = PatternFill("solid", fgColor="D0D0D0")
_ZEBRA_FILL = PatternFill("solid", fgColor="F5F5F5")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD = Font(bold=True)
_SECTION_FONT = Font(bold=True, color="1F3864")
_INT_FMT = "#,##0"
_CENTER = Alignment(horizontal="center")

_RECORDS_COLS = ["Дата", "Тип", "Сумма", "Категория", "Описание", "Счёт"]


def _hdr_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _CENTER


def _autowidth(ws) -> None:
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 45)


def _write_records_sheet(writer: pd.ExcelWriter, rows: list[dict]) -> None:
    df = (
        pd.DataFrame(rows, columns=_RECORDS_COLS)
        if rows
        else pd.DataFrame(columns=_RECORDS_COLS)
    )
    df.to_excel(writer, sheet_name="Записи", index=False)
    ws = writer.sheets["Записи"]
    _hdr_row(ws, 1, len(_RECORDS_COLS))
    for i, r in enumerate(rows, start=2):
        fill = _INCOME_FILL if r["Тип"] == "Доход" else _EXPENSE_FILL
        for c in range(1, len(_RECORDS_COLS) + 1):
            ws.cell(row=i, column=c).fill = fill
        ws.cell(row=i, column=3).number_format = _INT_FMT
    _autowidth(ws)


def _write_summary_sheet(writer: pd.ExcelWriter, summary: dict) -> None:
    ws = writer.book.create_sheet("Итоги")
    row = 1

    def sec(title: str) -> None:
        nonlocal row
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
        row += 1

    def kv(label: str, value, is_money: bool = False, bold: bool = False) -> None:
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        if bold:
            c1.font = _BOLD
            c2.font = _BOLD
        if is_money and isinstance(value, (int, float)):
            c2.number_format = _INT_FMT
        row += 1

    sec("Общая статистика")
    kv("Доходы", summary["total_income"], is_money=True, bold=True)
    kv("Расходы", summary["total_expense"], is_money=True, bold=True)
    kv("Баланс", summary["balance"], is_money=True, bold=True)
    kv("Транзакций (доход)", summary.get("count_income", 0))
    kv("Транзакций (расход)", summary.get("count_expense", 0))
    row += 1

    sec("Дополнительно")
    kv("Макс. разовый расход", summary.get("max_expense", 0), is_money=True)
    kv("Ср. расход в день", summary.get("avg_daily_expense", 0), is_money=True)
    kv("Ср. доход в день", summary.get("avg_daily_income", 0), is_money=True)
    fd, ld = summary.get("first_date"), summary.get("last_date")
    if fd and ld:
        kv("Период", f"{fd} — {ld}")
    row += 1

    for title, items in [
        ("Топ-5 расходов", summary.get("top5_expense", [])),
        ("Топ-5 доходов", summary.get("top5_income", [])),
    ]:
        if not items:
            continue
        sec(title)
        for c, h in enumerate(["Категория", "Сумма", "%"], start=1):
            ws.cell(row=row, column=c, value=h).font = _BOLD
        row += 1
        for cat, amt, pct in items:
            ws.cell(row=row, column=1, value=cat)
            c2 = ws.cell(row=row, column=2, value=amt)
            c2.number_format = _INT_FMT
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10


def _write_monthly_sheet(writer: pd.ExcelWriter, rows: list[dict]) -> None:
    monthly: dict[str, dict] = {}
    for r in rows:
        try:
            dt = datetime.strptime(r["Дата"], "%d.%m.%Y")
        except ValueError:
            continue
        key = dt.strftime("%Y-%m")
        if key not in monthly:
            monthly[key] = {
                "label": dt.strftime("%m.%Y"),
                "income": 0.0,
                "expense": 0.0,
                "count": 0,
            }
        monthly[key]["count"] += 1
        if r["Тип"] == "Доход":
            monthly[key]["income"] += r["Сумма"]
        else:
            monthly[key]["expense"] += r["Сумма"]

    if not monthly:
        return

    sorted_items = sorted(monthly.items())
    result = [
        {
            "Месяц": v["label"],
            "Доходы": v["income"],
            "Расходы": v["expense"],
            "Баланс": v["income"] - v["expense"],
            "Транзакций": v["count"],
        }
        for _, v in sorted_items
    ]
    result.append(
        {
            "Месяц": "ИТОГО",
            "Доходы": sum(v["income"] for _, v in sorted_items),
            "Расходы": sum(v["expense"] for _, v in sorted_items),
            "Баланс": sum(v["income"] - v["expense"] for _, v in sorted_items),
            "Транзакций": sum(v["count"] for _, v in sorted_items),
        }
    )

    df = pd.DataFrame(
        result, columns=["Месяц", "Доходы", "Расходы", "Баланс", "Транзакций"]
    )
    df.to_excel(writer, sheet_name="По месяцам", index=False)
    ws = writer.sheets["По месяцам"]
    _hdr_row(ws, 1, 5)

    n_data = len(result) - 1
    for i in range(2, n_data + 2):
        if i % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=i, column=c).fill = _ZEBRA_FILL
        for c in [2, 3, 4]:
            ws.cell(row=i, column=c).number_format = _INT_FMT

    totals_row = n_data + 2
    for c in range(1, 6):
        cell = ws.cell(row=totals_row, column=c)
        cell.fill = _TOTAL_FILL
        cell.font = _BOLD
    for c in [2, 3, 4]:
        ws.cell(row=totals_row, column=c).number_format = _INT_FMT

    _autowidth(ws)


def _write_categories_sheet(
    writer: pd.ExcelWriter, rows: list[dict], summary: dict
) -> None:
    cats: dict = {}
    for r in rows:
        key = (r["Категория"], r["Тип"])
        if key not in cats:
            cats[key] = {"amount": 0.0, "count": 0}
        cats[key]["amount"] += r["Сумма"]
        cats[key]["count"] += 1

    if not cats:
        return

    total_inc = summary.get("total_income") or 1
    total_exp = summary.get("total_expense") or 1

    result = sorted(
        [
            {
                "Категория": cat,
                "Тип": typ,
                "Сумма": d["amount"],
                "%": round(
                    d["amount"] / (total_inc if typ == "Доход" else total_exp) * 100, 1
                ),
                "Транзакций": d["count"],
            }
            for (cat, typ), d in cats.items()
        ],
        key=lambda x: -x["Сумма"],
    )

    df = pd.DataFrame(result, columns=["Категория", "Тип", "Сумма", "%", "Транзакций"])
    df.to_excel(writer, sheet_name="По категориям", index=False)
    ws = writer.sheets["По категориям"]
    _hdr_row(ws, 1, 5)

    for i, r in enumerate(result, start=2):
        fill = _INCOME_FILL if r["Тип"] == "Доход" else _EXPENSE_FILL
        for c in range(1, 6):
            ws.cell(row=i, column=c).fill = fill
        ws.cell(row=i, column=3).number_format = _INT_FMT

    _autowidth(ws)


def _build_export_sync(rows: list[dict], summary: dict) -> BytesIO:
    """Build styled Excel file with analytics sheets. CPU-bound, run in executor."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write_records_sheet(writer, rows)
        _write_summary_sheet(writer, summary)
        if rows:
            _write_monthly_sheet(writer, rows)
            _write_categories_sheet(writer, rows, summary)
    buf.seek(0)
    return buf


def _build_backup_sync(
    records_rows: list[dict],
    budgets: list[dict],
    snapshot_items: list[dict],
    wealth_items: list[dict],
) -> BytesIO:
    """Build full backup Excel. Reuses records sheet from _build_export_sync."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if records_rows:
            df = pd.DataFrame(records_rows, columns=_RECORDS_COLS)
        else:
            df = pd.DataFrame(columns=_RECORDS_COLS)
        df.to_excel(writer, sheet_name="Записи", index=False)

        if budgets:
            pd.DataFrame(budgets, columns=["Категория", "Лимит", "Активен"]).to_excel(
                writer, sheet_name="Бюджеты", index=False
            )

        if snapshot_items:
            pd.DataFrame(snapshot_items, columns=["Название", "Сумма"]).to_excel(
                writer, sheet_name="Накопления", index=False
            )

        if wealth_items:
            pd.DataFrame(
                wealth_items, columns=["Тип", "Название", "Сумма", "Заметка"]
            ).to_excel(writer, sheet_name="Активы", index=False)

    buf.seek(0)
    return buf


def _build_template_sync() -> BytesIO:
    """Build import template with headers and one example row."""
    buf = BytesIO()
    df = pd.DataFrame(
        [
            {
                "Дата": "01.01.2025",
                "Тип": "Расход",
                "Сумма": 500.0,
                "Категория": "Еда",
                "Описание": "Обед с коллегами",
                "Счёт": "Наличные",
            }
        ]
    )
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Импорт", index=False)
    buf.seek(0)
    return buf


def validate_import_row(row: dict, row_num: int) -> tuple[dict | None, str | None]:
    """Validate single row. Returns (parsed, None) or (None, error_str)."""
    # --- Дата ---
    raw_date = row.get("Дата")
    parsed_date: date_type | None = None
    if raw_date is not None:
        raw_str = str(raw_date).strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                parsed_date = datetime.strptime(raw_str, fmt).date()
                break
            except ValueError:
                continue
    if parsed_date is None:
        return None, f"Строка {row_num}: неверный формат даты «{raw_date}»"

    # --- Тип ---
    raw_op = str(row.get("Тип", "")).strip()
    if raw_op.lower() in ("расход", "-"):
        operation = "-"
    elif raw_op.lower() in ("доход", "+"):
        operation = "+"
    else:
        return (
            None,
            f"Строка {row_num}: неверный тип «{raw_op}» (ожидается Доход/Расход)",
        )

    # --- Сумма ---
    raw_amount = row.get("Сумма")
    try:
        amount = Decimal(str(raw_amount))
        if amount <= 0 or amount > MAX_AMOUNT:
            raise ValueError
    except InvalidOperation, ValueError, TypeError:
        return None, f"Строка {row_num}: неверная сумма «{raw_amount}»"

    # --- Категория ---
    raw_cat = clean_text(str(row.get("Категория", "")))
    if not raw_cat:
        return None, f"Строка {row_num}: категория пустая"
    if len(raw_cat) > MAX_CATEGORY_LENGTH:
        return (
            None,
            f"Строка {row_num}: категория слишком длинная (макс {MAX_CATEGORY_LENGTH})",
        )
    category = raw_cat[0].upper() + raw_cat[1:]

    # --- Счёт (опционально) ---
    raw_acc = row.get("Счёт")
    account_name: str | None = None
    if raw_acc is not None:
        s = clean_text(str(raw_acc))
        if s and s != "nan":
            if len(s) > MAX_ACCOUNT_NAME_LENGTH:
                return None, (
                    f"Строка {row_num}: название счёта слишком длинное "
                    f"(макс {MAX_ACCOUNT_NAME_LENGTH})"
                )
            account_name = s

    # --- Описание (опционально) ---
    raw_desc = row.get("Описание")
    description: str | None = None
    if raw_desc is not None:
        s = clean_text(str(raw_desc))
        if s and s != "nan":
            description = s[:255]

    return {
        "date": parsed_date,
        "operation": operation,
        "amount": amount,
        "category": category,
        "account_name": account_name,
        "description": description,
    }, None


def parse_import_file(
    file_bytes: bytes,
    max_rows: int = 1000,
) -> tuple[list[dict], list[str], int]:
    """Parse xlsx bytes. Returns (valid_rows, error_strings, duplicates_placeholder=0).

    valid_rows keys: date, operation, amount, category, account_name.
    """
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, dtype=str)
    except Exception as e:
        return [], [f"Не удалось прочитать файл: {e}"], 0

    required = {"Дата", "Тип", "Сумма", "Категория"}
    missing = required - set(df.columns)
    if missing:
        return [], [f"Отсутствуют столбцы: {', '.join(missing)}"], 0

    # Drop fully empty rows
    df = df.dropna(how="all")

    if len(df) > max_rows:
        return [], [f"Файл содержит более {max_rows} строк. Разбейте на части."], 0

    valid_rows: list[dict] = []
    errors: list[str] = []

    for idx, row in enumerate(df.to_dict("records"), start=2):
        parsed, err = validate_import_row(row, idx)
        if err:
            errors.append(err)
        elif parsed is not None:
            valid_rows.append(parsed)

    return valid_rows, errors, 0
